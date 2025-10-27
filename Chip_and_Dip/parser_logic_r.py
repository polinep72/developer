# parser_logic.py
import datetime
import os
import random
import time
import warnings
import logging
import logging.handlers

import undetected_chromedriver as uc
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
import pytz
from decouple import config, UndefinedValueError

warnings.filterwarnings('ignore')

# --- Настройка логирования ---
logger = logging.getLogger(__name__)  # Logger будет настроен через setup_logging


def setup_logging(log_level_console=logging.INFO, log_level_file=logging.DEBUG):
    """Настраивает систему логирования."""
    global logger
    logger.setLevel(logging.DEBUG)  # Устанавливаем общий уровень DEBUG для логгера

    # Удаляем существующие обработчики, чтобы избежать дублирования при перезапусках (например, в Flask debug режиме)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    log_formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(module)s.%(funcName)s:%(lineno)d - %(message)s'
    )

    # Консольный обработчик
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)
    console_handler.setLevel(log_level_console)
    logger.addHandler(console_handler)

    # Файловый обработчик
    log_file_path = "app.log"  # Имя файла логов
    try:
        # Используем RotatingFileHandler для ограничения размера файла логов (опционально, но полезно)
        # file_handler = logging.handlers.RotatingFileHandler(
        #     log_file_path, maxBytes=5*1024*1024, backupCount=2, encoding='utf-8'
        # )
        # Или обычный FileHandler
        file_handler = logging.FileHandler(log_file_path, mode='a', encoding='utf-8')
        file_handler.setFormatter(log_formatter)
        file_handler.setLevel(log_level_file)
        logger.addHandler(file_handler)
        logger.info(
            f"Логирование настроено. Консоль: {logging.getLevelName(log_level_console)}, Файл ('{log_file_path}'): {logging.getLevelName(log_level_file)}")
    except Exception as e:
        # Если не удалось настроить файловый логгер, продолжаем с консольным
        logger.error(f"Не удалось настроить файловый логгер для '{log_file_path}': {e}",
                     exc_info=False)  # exc_info=False, т.к. уже есть console handler
        logger.info("Логирование будет осуществляться только в консоль.")


# --- Глобальные переменные конфигурации ---
# APP_CONFIG будет заполняться функцией load_parser_config или передан в run_parsing_process
APP_CONFIG = {}


def get_element_by_xpath(url):
    """
    Извлекает числовое значение с веб-страницы по заданному XPath.
    APP_CONFIG должен быть предварительно заполнен.
    """
    logger.debug(f"Запрос элемента по XPath для URL: {url}")
    if not url:
        logger.warning("URL не предоставлен для get_element_by_xpath.")
        return -1  # Ошибка: нет URL

    # --- ОБНОВЛЕННЫЙ XPATH (из вашего оригинального кода) ---
    xpath = "//span[contains(@class, 'item__avail') and contains(@class, 'item__avail_delivery')]/b"
    logger.debug(f"Используемый XPath: {xpath}")

    options = uc.ChromeOptions()

    try:
        headless_flag = bool(int(APP_CONFIG.get('headless', '0')))  # Получаем из APP_CONFIG
        options.headless = headless_flag
        logger.debug(f"Headless режим: {options.headless}")
    except ValueError:
        logger.warning(
            f"Некорректное значение для 'headless' в конфигурации: '{APP_CONFIG.get('headless')}'. Используется False.",
            exc_info=False)
        options.headless = False

    chrome_executable_path = APP_CONFIG.get('chrome_executable_path')  # Получаем из APP_CONFIG

    if chrome_executable_path and os.path.exists(chrome_executable_path):
        options.binary_location = chrome_executable_path
        logger.info(f"Используется Chrome из: {chrome_executable_path}")
    elif chrome_executable_path:
        logger.warning(
            f"Указанный путь к Chrome '{chrome_executable_path}' не существует. undetected_chromedriver попытается найти Chrome автоматически.")
    else:
        logger.info(
            "Путь к Chrome (chrome_executable_path) не указан в конфигурации. undetected_chromedriver попытается найти Chrome автоматически.")

    driver = None
    result = -1  # Значение по умолчанию в случае ошибки
    try:
        logger.debug(
            f"Инициализация uc.Chrome. Headless: {options.headless}, Binary Location: {options.binary_location or 'Автоматически'}")
        driver = uc.Chrome(options=options)
        logger.debug(f"Драйвер Chrome инициализирован. Загрузка URL: {url}")
        driver.get(url)
        logger.debug(f"URL '{url}' загружен.")

        wait_timeout = 15  # Секунд
        logger.debug(f"Ожидание элемента (до {wait_timeout} сек) по XPath: {xpath}")
        try:
            element = WebDriverWait(driver, wait_timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            logger.debug(f"Элемент по XPath '{xpath}' найден. Текст: '{element.text}'")

            raw_text = element.text
            cleaned_text = ''.join(filter(str.isdigit, raw_text))
            if cleaned_text:
                result = int(cleaned_text)
            else:
                logger.warning(f"Текст элемента '{raw_text}' не содержит цифр после очистки для URL: {url}")
                result = 0  # Если элемент найден, но текста нет или он не содержит цифр
            logger.info(f"Для URL '{url}' получен результат: {result}")

        except TimeoutException:
            logger.warning(
                f"Элемент по XPath '{xpath}' не найден в течение {wait_timeout} сек для URL: {url} (TimeoutException)")
            result = 0  # Если элемент не найден по таймауту, считаем остаток 0
        # NoSuchElementException теперь будет ловиться TimeoutException, если элемент не появится за время ожидания

    except TypeError as e:  # Например, если APP_CONFIG не настроен должным образом
        logger.error(
            f"TypeError в get_element_by_xpath при обработке URL '{url}'. Возможно, проблема с конфигурацией или путем к Chrome. Ошибка: {e}",
            exc_info=True)
        result = -1  # Ошибка
    except Exception as e:
        logger.error(f"Непредвиденная ошибка в get_element_by_xpath при обработке URL '{url}'. Ошибка: {e}",
                     exc_info=True)
        result = -1  # Ошибка
    finally:
        if driver:
            logger.debug(f"Закрытие драйвера Chrome для URL: {url}")
            driver.quit()
    return result


def run_parsing_process(path_to_xlsx_file, app_config_values, progress_callback):
    """
    Основная логика парсинга.
    path_to_xlsx_file: Путь к исходному Excel файлу.
    app_config_values: Словарь с конфигурацией, полученный из load_parser_config.
    progress_callback: Функция для обновления прогресса (current, total, message, file_path=None, error=None).
    Возвращает путь к обработанному файлу или None в случае ошибки.
    """
    global APP_CONFIG
    APP_CONFIG = app_config_values  # Устанавливаем глобальную конфигурацию для этого процесса

    logger.info(f"Начало обработки файла Excel (веб-запрос): {path_to_xlsx_file}")
    progress_callback(0, 1, "Инициализация и проверка файла...", None, None)

    if not os.path.exists(path_to_xlsx_file):
        error_msg = f"Файл «{path_to_xlsx_file}» не найден."
        logger.error(error_msg)
        progress_callback(1, 1, "Ошибка: Файл не найден", None, error_msg)
        return None

    try:
        workbook = openpyxl.load_workbook(path_to_xlsx_file)
        worksheet = workbook.active
        logger.debug(f"Файл '{path_to_xlsx_file}' успешно открыт. Лист: '{worksheet.title}'.")
    except Exception as e:
        error_msg = f"Ошибка при открытии или чтении файла Excel '{path_to_xlsx_file}'. Детали: {e}"
        logger.error(error_msg, exc_info=True)
        progress_callback(1, 1, "Ошибка открытия Excel", None, error_msg)
        return None

    n_col = 0
    for col_i in range(1, worksheet.max_column + 2):  # +2 чтобы найти следующий пустой столбец
        if worksheet.cell(row=1, column=col_i).value in ['', None]:
            n_col = col_i
            break
    if n_col == 0:  # Если все столбцы до max_column+1 заняты (маловероятно, но на всякий случай)
        n_col = worksheet.max_column + 1
    logger.debug(f"Данные будут записываться в столбец: {n_col}")

    time_format = APP_CONFIG.get('time_format', '%Y-%m-%d %H:%M:%S')
    try:
        current_time_str = datetime.datetime.now(pytz.timezone('Europe/Moscow')).strftime(time_format)
        worksheet.cell(row=1, column=n_col).value = current_time_str
        logger.info(f"В ячейку (1, {n_col}) записана временная метка: {current_time_str}")
    except Exception as e:
        logger.warning(f"Не удалось записать временную метку: {e}")
        # Продолжаем работу, даже если метку не удалось записать

    rows_to_process = []
    # Собираем все строки, где в 3-м столбце есть что-то похожее на URL или гиперссылку
    for row_idx in range(2, worksheet.max_row + 1):
        cell_val = worksheet.cell(row=row_idx, column=3).value
        cell_link = worksheet.cell(row=row_idx, column=3).hyperlink
        if (isinstance(cell_val, str) and cell_val.strip()) or (cell_link and cell_link.target):
            rows_to_process.append(row_idx)

    total_rows_to_parse = len(rows_to_process)
    if total_rows_to_parse == 0:
        message = "Нет URL для обработки в файле."
        logger.info(message)
        progress_callback(1, 1, message, None, None)  # Завершаем с сообщением
        # Сохраняем файл с меткой времени, если она была добавлена
        try:
            workbook.save(path_to_xlsx_file)  # Перезаписываем исходный
            logger.info(f"Файл '{path_to_xlsx_file}' сохранен (нет URL для обработки).")
            return path_to_xlsx_file  # Возвращаем путь к исходному файлу
        except Exception as e:
            error_msg = f"Ошибка при сохранении файла '{path_to_xlsx_file}' (нет URL): {e}"
            logger.error(error_msg, exc_info=True)
            progress_callback(1, 1, "Ошибка сохранения файла", None, error_msg)
            return None

    for i, row_i in enumerate(rows_to_process):
        current_step = i + 1
        progress_callback(
            current_step,
            total_rows_to_parse,
            f"Обработка строки {row_i} ({current_step}/{total_rows_to_parse})...",
            None,
            None
        )
        logger.debug(f"Обработка строки Excel: {row_i}")

        url_cell = worksheet.cell(row=row_i, column=3)
        url_value = url_cell.value
        url_to_process = None

        if url_cell.hyperlink and url_cell.hyperlink.target:
            url_to_process = url_cell.hyperlink.target
            logger.debug(f"Строка {row_i}: URL извлечен из гиперссылки: {url_to_process}")
        elif isinstance(url_value, str) and url_value.strip():
            url_to_process = url_value.strip()
            logger.debug(f"Строка {row_i}: URL извлечен из значения ячейки: {url_to_process}")

        if not url_to_process or not url_to_process.lower().startswith(('http://', 'https://')):
            logger.warning(f"Строка {row_i}, столбец 3: некорректный или отсутствующий URL '{url_value}'. Пропускаем.")
            worksheet.cell(row=row_i, column=n_col).value = 'Некорректный URL'
            # Короткая задержка для обновления UI через SSE, если много таких подряд
            time.sleep(0.01)
            continue

        logger.info(f"Строка {row_i}: Обработка URL: {url_to_process}")
        stock = get_element_by_xpath(url_to_process)  # APP_CONFIG используется внутри get_element_by_xpath

        if stock == -1:  # Ошибка при парсинге
            logger.error(f"Строка {row_i}: Ошибка при получении данных для URL: {url_to_process}")
            worksheet.cell(row=row_i, column=n_col).value = 'Ошибка парсинга'
        else:  # stock может быть 0 или положительным числом
            logger.info(
                f"Строка {row_i}: Для URL '{url_to_process}' получен сток: {stock}. Запись в ячейку ({row_i}, {n_col}).")
            worksheet.cell(row=row_i, column=n_col).value = stock
            if APP_CONFIG.get('console_log_active', False):  # Логирование для Flask-консоли
                logger.info(f"(Веб-процесс) URL: {url_to_process}, Сток: {stock}")

        # Случайная задержка между запросами к сайту
        delay = random.uniform(0.2, 1.5)  # Немного увеличил диапазон
        logger.debug(f"Задержка перед следующим URL: {delay:.2f} сек.")
        time.sleep(delay)

    # --- Сохранение файла ---
    # Вариант: перезаписываем исходный файл
    output_filepath = path_to_xlsx_file

    # Вариант: создаем новый файл с временной меткой (предпочтительнее для веб, чтобы не было конфликтов)
    # original_dir = os.path.dirname(path_to_xlsx_file)
    # original_basename = os.path.basename(path_to_xlsx_file)
    # name_part, ext_part = os.path.splitext(original_basename)
    # timestamp_save = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    # output_filename = f"{name_part}_parsed_{timestamp_save}{ext_part}"
    # output_filepath = os.path.join(original_dir, output_filename)

    try:
        workbook.save(output_filepath)
        logger.info(f"Файл Excel '{output_filepath}' успешно сохранен.")
        progress_callback(
            total_rows_to_parse,
            total_rows_to_parse,
            "Обработка завершена! Файл сохранен.",
            output_filepath,  # Передаем путь к файлу
            None
        )
        return output_filepath
    except Exception as e:
        error_msg = f"Ошибка при сохранении файла Excel '{output_filepath}'. Детали: {e}"
        logger.error(error_msg, exc_info=True)
        progress_callback(
            total_rows_to_parse,
            total_rows_to_parse,
            "Ошибка сохранения файла",
            None,
            error_msg
        )
        return None


def load_parser_config():
    """
    Загружает конфигурацию из .env файла.
    Возвращает словарь с конфигурацией или выбрасывает исключение.
    """
    global APP_CONFIG
    logger.info("Загрузка конфигурации парсера...")
    config_data = {}
    try:
        # Используем импортированный 'config' напрямую
        config_data['xlsx_file'] = config('XLSX_FILE')
        config_data['headless'] = config('HEADLESS_MODE', default='0')
        config_data['time_format'] = config('TIME_FORMAT', default='%Y-%m-%d %H:%M:%S')
        config_data['console_log_active'] = config('CONSOLE_LOG_ACTIVE', default='0', cast=bool)
        config_data['chrome_executable_path'] = config('CHROME_EXECUTABLE_PATH', default=None)

        # Определяем абсолютный путь к Excel файлу
        # Директория, где лежит parser_logic.py (и, предположительно, .env файл)
        # config() ищет .env файл, начиная с директории запускаемого скрипта и поднимаясь вверх.
        # Если app.py и parser_logic.py в одной папке, и .env там же, то все будет найдено.

        script_dir_or_project_root = os.path.dirname(os.path.abspath(__file__))  # Директория parser_logic.py
        # Если ваш .env файл лежит в той же директории, что и parser_logic.py (или выше),
        # то decouple его найдет автоматически.

        configured_xlsx_path_from_env = config_data['xlsx_file']

        if os.path.isabs(configured_xlsx_path_from_env):
            config_data['actual_xlsx_file_path'] = configured_xlsx_path_from_env
        else:
            # Если путь относительный, считаем его относительно директории, где лежит .env файл
            # (которую decouple находит) или, для большей предсказуемости,
            # относительно директории, где лежит parser_logic.py.
            # Для простоты, если XLSX_FILE="data.xlsx", он будет искаться рядом с parser_logic.py
            # или в директории, где decouple нашел .env.
            # Лучше всего, если .env и скрипты в одной папке, а data.xlsx там же или путь указан относительно этой папки.
            config_data['actual_xlsx_file_path'] = os.path.join(script_dir_or_project_root,
                                                                configured_xlsx_path_from_env)

        logger.info(f"Конфигурация парсера загружена: {config_data}")
        APP_CONFIG = config_data
        return config_data
    except UndefinedValueError as e:
        logger.critical(f"Критическая ошибка: не найдена обязательная переменная конфигурации в .env файле: {e}")
        raise
    except Exception as e:
        logger.critical(f"Критическая ошибка при загрузке конфигурации парсера: {e}", exc_info=True)
        raise


# Блок для возможности отдельного запуска (например, для отладки самого парсера)
if __name__ == '__main__':
    setup_logging(log_level_console=logging.DEBUG,
                  log_level_file=logging.DEBUG)  # Более детальное логирование для отладки

    logger.info("Запуск parser_logic.py в автономном режиме для отладки...")

    try:
        cfg = load_parser_config()  # Загружаем конфигурацию

        if cfg and cfg.get('actual_xlsx_file_path'):
            xlsx_file_to_test = cfg['actual_xlsx_file_path']
            logger.info(f"Тестирование парсера с файлом: {xlsx_file_to_test}")

            if not os.path.exists(xlsx_file_to_test):
                logger.error(
                    f"Файл для теста '{xlsx_file_to_test}' не найден! Проверьте путь в .env и его расположение.")
            else:
                # Пример callback-функции для консольного режима
                def console_progress_callback(current, total, message, file_path=None, error_msg=None):
                    progress_percent = (current / total) * 100 if total > 0 else 0
                    log_message = f"Прогресс: {current}/{total} ({progress_percent:.2f}%) - {message}"
                    if file_path:
                        log_message += f" | Файл: {file_path}"
                    if error_msg:
                        log_message += f" | Ошибка: {error_msg}"
                    logger.info(log_message)  # Используем логгер вместо print


                result_file = run_parsing_process(xlsx_file_to_test, cfg, console_progress_callback)

                if result_file:
                    logger.info(f"Автономный запуск парсера завершен. Результат в файле: {result_file}")
                else:
                    logger.error("Автономный запуск парсера завершился без сохранения файла или с ошибкой.")
        else:
            logger.error("Не удалось загрузить конфигурацию или определить путь к файлу Excel для автономного запуска.")

    except Exception as e:
        logger.critical(f"Ошибка во время автономного запуска parser_logic.py: {e}", exc_info=True)

    logger.info("Автономный режим parser_logic.py завершен.")