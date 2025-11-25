"""
Экспорт данных в Excel (XLSX)
"""
import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def export_dashboard_excel(
    payload: Dict[str, Any],
    filters: Dict[str, Any],
    start_date: date,
    end_date: date,
) -> bytes:
    """Экспорт данных дашборда в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика бронирований"
    
    # Стили
    header_fill = PatternFill(start_color="027BBC", end_color="027BBC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    row = 1
    
    # Заголовок
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Аналитика бронирований рабочих мест"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    row += 2
    
    # Параметры фильтрации
    ws[f'A{row}'] = "Параметры фильтрации"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Начало периода"
    ws[f'B{row}'] = start_date.strftime("%d.%m.%Y")
    row += 1
    
    ws[f'A{row}'] = "Конец периода"
    ws[f'B{row}'] = end_date.strftime("%d.%m.%Y")
    row += 1
    
    equipment_list = filters.get("equipment", [])
    ws[f'A{row}'] = "Оборудование"
    ws[f'B{row}'] = ", ".join(equipment_list) if equipment_list else "Все"
    row += 1
    
    ws[f'A{row}'] = "Целевая загрузка (ч/день)"
    ws[f'B{row}'] = filters.get("target_load", 8)
    row += 1
    
    ws[f'A{row}'] = "Фактическая загрузка"
    ws[f'B{row}'] = payload.get("utilization", "0%")
    row += 2
    
    # Суммарная наработка по оборудованию
    ws[f'A{row}'] = "Суммарная наработка по оборудованию"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Прибор"
    ws[f'B{row}'] = "Часы"
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    equipment_summary = payload.get("equipmentSummary", [])
    for item in equipment_summary:
        ws[f'A{row}'] = item.get("name", "")
        ws[f'B{row}'] = item.get("hours", 0)
        ws[f'B{row}'].number_format = "0.00"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.border = border
        row += 1
    
    row += 1
    
    # Активность пользователей
    ws[f'A{row}'] = "Активность пользователей"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Пользователь"
    ws[f'B{row}'] = "Часы"
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    users = payload.get("users", [])
    for item in users:
        ws[f'A{row}'] = item.get("name", "")
        ws[f'B{row}'] = item.get("hours", 0)
        ws[f'B{row}'].number_format = "0.00"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.border = border
        row += 1
    
    # Автоматическая ширина столбцов
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    # Сохранение в байты
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_equipment_excel(
    equipment_data: List[Dict[str, Any]],
    equipment_type: str,
) -> bytes:
    """Экспорт данных модуля СИ в Excel"""
    wb = Workbook()
    ws = wb.active
    
    # Название листа в зависимости от типа
    type_names = {
        "si": "Средства измерения",
        "io": "Испытательное оборудование",
        "vo": "Вспомогательное оборудование",
        "gosregister": "Госреестр"
    }
    ws.title = type_names.get(equipment_type, "Оборудование")
    
    # Стили
    header_fill = PatternFill(start_color="027BBC", end_color="027BBC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    if not equipment_data:
        ws['A1'] = "Нет данных для экспорта"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    # Определяем заголовки на основе первого элемента
    first_item = equipment_data[0]
    headers = list(first_item.keys())
    
    # Заголовки
    row = 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Данные
    for item in equipment_data:
        row += 1
        for col_idx, header in enumerate(headers, start=1):
            value = item.get(header, "")
            # Обработка специальных типов
            if isinstance(value, (date, datetime)):
                value = value.strftime("%d.%m.%Y")
            elif value is None:
                value = ""
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
    
    # Автоматическая ширина столбцов
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Сохранение в байты
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_bookings_excel(
    csv_content: str,
    target_date: date,
    include_all: bool = False,
) -> bytes:
    """Экспорт бронирований в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Бронирования"
    
    # Стили
    header_fill = PatternFill(start_color="027BBC", end_color="027BBC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    row = 1
    
    # Заголовок
    ws.merge_cells(f'A{row}:H{row}')
    title = f"Бронирования за {target_date.strftime('%d.%m.%Y')}"
    ws[f'A{row}'] = title
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    row += 2
    
    # Определяем заголовки (приводим к реальному набору данных CSV)
    csv_lines = csv_content.strip().split("\n")
    if len(csv_lines) < 2:
        ws.cell(row=row, column=1, value="Нет данных для экспорта")
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    csv_headers = [h.strip() for h in csv_lines[0].split(";")]
    has_user_column = include_all and "Пользователь" in csv_headers

    headers = [
        "Дата",
        "Время начала",
        "Время окончания",
        "Интервал",
        "Длительность (мин)",
        "Оборудование",
        "Категория",
    ]
    if has_user_column:
        headers.append("Пользователь")
    headers.append("Статус")
    
    # Заголовки таблицы
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    row += 1
    
    # Пропускаем заголовок (уже добавлен выше) и читаем данные
    for line in csv_lines[1:]:
        if not line.strip():
            continue
        values = [v.strip() for v in line.split(";")]
        row_data = {csv_headers[i]: values[i] for i in range(min(len(csv_headers), len(values)))}
        
        duration_minutes = ""
        duration_hours = row_data.get("Длительность (ч)")
        if duration_hours:
            try:
                duration_minutes = str(int(float(duration_hours) * 60))
            except ValueError:
                duration_minutes = duration_hours
        
        excel_row = [
            row_data.get("Дата", ""),
            row_data.get("Время начала", ""),
            row_data.get("Время окончания", ""),
            row_data.get("Интервал", ""),
            duration_minutes,
            row_data.get("Оборудование", ""),
            row_data.get("Категория", ""),
        ]
        if has_user_column:
            excel_row.append(row_data.get("Пользователь", ""))
        excel_row.append(row_data.get("Статус", ""))
        
        for col_idx, value in enumerate(excel_row, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
        row += 1
    
    # Автоматическая ширина столбцов
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Сохранение в байты
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

